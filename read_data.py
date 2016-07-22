from openpyxl import load_workbook
import pandas as pd
from pandas import read_pickle
import warnings
import os
import sys
import re


def preprocess_string(ori_str):
    ori_str = str(ori_str).replace('*', '').replace(':', '').replace('\'\'', '').replace('_', '').replace('-', '')
    ori_str = ori_str.replace('/', ' ').replace('.', ' ').replace('|', ' ')
    ori_str = del_space_ahead(ori_str)

    ori_str = re.sub(r'[^a-zA-Z0-9 ]+', '', ori_str)
    return ori_str.lower()


def del_space_ahead(ori_str):
    count = 0
    for iChar in range(len(ori_str)):
        if ori_str[iChar] == ' ':
            count += 1
        else:
            break
    return ori_str.replace(' ', '', count)


def get_data_from_file(file_name, _get_ac):
    try:
        warnings.simplefilter("ignore")
        wb = load_workbook(filename=file_name, read_only=True)
        w_data = wb['data']
        max_row = w_data.max_row

        r_row = 18
        if w_data._get_cell(1, r_row).value != 'Cm Motif Dep':
            print '[Debugging]: R columns name is not match.'
            raise Exception

        r_data = []
        gen_r = w_data.get_squared_range(r_row, 2, r_row, max_row)
        for iRow in gen_r:
            if iRow[0].value is not None:
                r_data.append(preprocess_string(iRow[0].value))

        if _get_ac:
            ac_row = 29
            if w_data._get_cell(1, ac_row).value != 'Reason of Removal':
                print '[Debugging]: AC columns name is not match. '
                raise Exception

            ac_data = []
            gen_ac = w_data.get_squared_range(ac_row, 2, ac_row, max_row)
            for iRow in gen_ac:
                if iRow[0].value is not None:
                    ac_data.append(str(iRow[0].value))

            data_re = pd.DataFrame({
                'r': r_data,
                'ac': ac_data
            })
            return True, data_re

        else:
            data_re = pd.DataFrame({
                'r': r_data
            })
            return True, data_re

    except:
        print '[Debugging]: Error in reading data from xlsm file: ' + file_name
        return False, []


def get_data_r_ac(data_folder, get_ac, read_cache=True):

    if read_cache and os.path.exists('./cache/data_r_ac.pkl'):
        print 'Load data from cache...'
        return read_pickle('./cache/data_r_ac.pkl')
    else:
        frames = []
        count = 1
        for iFile in os.listdir(data_folder):
            if iFile.endswith(".xlsm"):
                suc, data_pd = get_data_from_file(os.path.join(data_folder, iFile), get_ac)
                if suc:
                    frames.append(data_pd)
                    print '[Status]: Finish reading xlsm file ', count
                    count += 1

        if len(frames)>0:
            df = pd.concat(frames)
            df.to_pickle('./cache/data_r_ac.pkl')
            return df
        else:
            print '[Debugging]: Get no data!'
            sys.exit(1)
